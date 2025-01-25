import logging
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl.reader.excel import load_workbook

# Configurazione del logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()


def concat_fields(fields):
    """
    Funzione che concatena una lista di campi in un'unica stringa separata da ' - '.

    Parametri:
    fields (list): Lista di stringhe da concatenare.

    Ritorna:
    str: Una stringa risultante dalla concatenazione dei campi.
    """
    logger.info("Concatenando i campi CSV.")
    return ' - '.join(fields)


def load_processed_files(checkpoint_path):
    """
    Carica l'elenco dei file già elaborati da uno o più file di checkpoint.

    Parameters:
        checkpoint_path (str): Percorso al file di checkpoint o alla cartella contenente file di checkpoint.

    Returns:
        set: Insieme dei nomi dei file già elaborati.
    """
    processed_files = set()

    if os.path.isdir(checkpoint_path):
        logger.info(f"Caricamento file elaborati da tutti i file di checkpoint nella cartella: {checkpoint_path}")
        # Scansiona tutti i file nella cartella
        for root, _, files in os.walk(checkpoint_path):
            for file in files:
                file_path = os.path.join(root, file)
                if os.path.isfile(file_path):
                    with open(file_path, 'r') as f:
                        file_data = set(f.read().splitlines())
                        processed_files.update(file_data)
                        logger.info(f"File di checkpoint caricato: {file} ({len(file_data)} file trovati).")
    elif os.path.isfile(checkpoint_path):
        logger.info(f"Caricamento file elaborati dal file di checkpoint: {checkpoint_path}")
        with open(checkpoint_path, 'r') as f:
            processed_files.update(f.read().splitlines())
            logger.info(f"File elaborati caricati: {len(processed_files)} file trovati.")
    else:
        logger.warning(f"Nessun file o cartella di checkpoint trovata: {checkpoint_path}. Restituisco un set vuoto.")

    return processed_files


def update_checkpoint(checkpoint_file, new_files):
    """
    Aggiorna il file di checkpoint aggiungendo i nuovi file elaborati.

    Parameters:
        checkpoint_file (str): Percorso del file di checkpoint.
        new_files: Insieme dei nuovi file elaborati da aggiungere.
    """
    logger.info(f"Aggiornamento del checkpoint con {len(new_files)} nuovi file.")
    with open(checkpoint_file, 'a') as f:
        for file in new_files:
            f.write(file + "\n")
    logger.info(f"Checkpoint aggiornato con successo: {checkpoint_file}")


def write_excel_from_df_in_append_mode(file_path, new_data):
    """
    Salva i dati nel file Excel in modalità append.

    Parametri:
    file_path (str): Percorso del file Excel.
    new_data (DataFrame): Il nuovo DataFrame da aggiungere.
    """
    if not os.path.exists(file_path):
        # Se il file non esiste, crea un nuovo file con i dati
        new_data.to_excel(file_path, index=False, engine='openpyxl')
        logger.info(f"File Excel creato: {file_path}")
    else:
        # Carica il file Excel esistente
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Ottieni l'indice della prima riga vuota
        start_row = sheet.max_row + 1

        # Scrivi i nuovi dati sotto quelli esistenti
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            new_data.to_excel(writer, index=False, header=False, startrow=start_row - 1)
        logger.info(f"Nuovi dati aggiunti al file Excel: {file_path}")


def extract_date_from_filename(filename):
    """
    Estrae una data dal nome del file in formato 'gennaio2025' e la converte in '01-01-2025'.

    Parametri:
    - filename (str): Nome del file.

    Ritorna:
    - str: Data in formato 'dd-mm-yyyy'.
    """
    try:
        # Mappa dei mesi in italiano
        mesi = {
            "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
            "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
            "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12
        }

        # Estrarre il nome del mese e l'anno
        for mese, numero in mesi.items():
            if mese in filename.lower():
                anno = re.search(r'\d{4}', filename).group()
                data = datetime(int(anno), numero, 1)  # Primo giorno del mese
                return data.strftime("%d-%m-%Y")

        logger.warning(f"Impossibile estrarre una data valida dal file: {filename}")
        return None
    except Exception as e:
        logger.error(f"Errore durante l'estrazione della data dal file {filename}: {e}")
        return None


def upsert_to_csv(dataframe, csv_path, key_column):
    """
    Scrive un DataFrame in modalità upsert in un file CSV.
    Se la chiave esiste già, il record viene ignorato.

    :param dataframe: DataFrame da scrivere.
    :param csv_path: Percorso del file CSV.
    :param key_column: Nome della colonna chiave per l'unicità.
    """
    if not os.path.exists(csv_path):
        # Se il file non esiste, crea il CSV con i dati del DataFrame
        dataframe.to_csv(csv_path, index=False)
        logger.info(f"File creato: {csv_path}")
        return

    # Leggi i dati esistenti dal CSV
    existing_data = pd.read_csv(csv_path, delimiter=';', header=0)

    # Trova i record nuovi (che non sono presenti nel CSV esistente)
    new_data = dataframe[~dataframe[key_column].isin(existing_data[key_column])]

    if not new_data.empty:
        # Aggiungi solo i nuovi record al CSV
        updated_data = pd.concat([existing_data, new_data], ignore_index=True)
        updated_data.to_csv(csv_path, index=False)
        logger.info(f"Aggiunti {len(new_data)} record nuovi a {csv_path}")
    else:
        logger.info("Nessun nuovo record da aggiungere.")

def print_project_structure(directory, exclude_folders=None, indent=0, last_item=False):
    # Imposta una lista di cartelle da escludere, se non viene specificata ne usa una vuota
    exclude_folders = exclude_folders or []

    # Aggiungi '_pycache_' alla lista delle cartelle da escludere
    exclude_folders.append("__pycache__")

    # Lista tutti gli elementi nella directory
    items = os.listdir(directory)
    for i, item in enumerate(items):
        item_path = os.path.join(directory, item)

        # Se l'elemento è una cartella da escludere, salta l'iterazione
        if item in exclude_folders:
            continue

        # Verifica se l'elemento è una directory
        if os.path.isdir(item_path):
            # Verifica se è l'ultimo elemento della directory per decidere il simbolo da usare
            is_last_item = (i == len(items) - 1)

            # Aggiungi simboli di struttura per la leggibilità
            connector = "└── " if last_item else "├── "
            line = '  ' * indent + (connector if is_last_item else "├── ") + item
            print(line)

            # Chiamata ricorsiva per la directory
            print_project_structure(item_path, exclude_folders, indent + 1, is_last_item)

        # Includi solo i file .py
        elif item.endswith(".py"):
            is_last_item = (i == len(items) - 1)
            connector = "└── " if last_item else "├── "
            line = '  ' * indent + (connector if is_last_item else "├── ") + item
            print(line)
