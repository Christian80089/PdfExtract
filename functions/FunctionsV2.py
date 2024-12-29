import logging
import os
import time

import pandas as pd
import pdfplumber
import psycopg2
import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait

# Configurazione del logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()


def concat_fields(fields):
    """
    Funzione che concatena una lista di campi in un'unica stringa separata da ' - '.

    Parametri:
    fields (list): Lista di stringhe da concatenare.

    Ritorna:
    str: Una stringa risultante dalla concatenazione dei campi.
    """
    logger.info("Concatenando i campi CSV.")
    return ' - '.join(fields)


def extract_pdf_data(pdf_path):
    """
    Estrae il testo da un file PDF specificato.

    Parametri:
    pdf_path (str): Il percorso del file PDF da cui estrarre il testo.

    Ritorna:
    str: Il testo estratto dal PDF. In caso di errore, restituisce una stringa vuota.
    """
    logger.info(f"Inizio estrazione dati dal PDF: {pdf_path}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Estrai il testo da tutte le pagine del PDF e concatenalo in un'unica stringa
            full_text = ''.join([page.extract_text() for page in pdf.pages])
        logger.info(f"Estrazione completata, {len(full_text)} caratteri estratti.")
        return full_text
    except Exception as e:
        logger.error(f"Errore durante l'estrazione del PDF: {e}")
        return ""  # Restituisce una stringa vuota in caso di errore


def run_copilot(input_text, info_to_extract):
    """
    Automazione dell'interazione con il sito Copilot di Microsoft con logica di fallback.

    Parametri:
    input_text (str): Il testo non formattato che verrà analizzato da Copilot.

    Questo metodo automatizza l'interazione con Copilot: apre il sito, invia richieste e riceve risposte.
    In caso di errori, vengono loggati e il browser viene chiuso correttamente.
    """
    logger.info("Inizio interazione con Copilot.")

    driver = None
    try:
        # Avvia il browser Chrome
        driver = webdriver.Chrome()
        logger.info("Caricamento della pagina Copilot.")
        driver.get("https://copilot.microsoft.com/onboarding")

        # Tentativo di cliccare su "Inizia"
        try:
            logger.info("Tentativo di cliccare sul pulsante 'Inizia'.")
            WebDriverWait(driver, 5).until(
                ec.element_to_be_clickable((By.XPATH, '//*[@title="Inizia"]'))
            ).click()
            logger.info("Pulsante 'Inizia' cliccato.")
        except Exception:
            logger.warning("Pulsante 'Inizia' non trovato. Procedo al passaggio successivo.")

        # Tentativo di inserire "Christian"
        try:
            logger.info("Tentativo di inserire 'Christian' nella casella di input.")
            WebDriverWait(driver, 5).until(
                ec.presence_of_element_located((By.ID, "userInput"))
            ).send_keys("Christian", Keys.RETURN)
            logger.info("'Christian' inserito con successo.")
        except Exception:
            logger.warning("Casella di input per 'Christian' non trovata. Procedo al passaggio successivo.")

        # Tentativo di cliccare su "Avanti"
        try:
            logger.info("Tentativo di cliccare sul pulsante 'Avanti'.")
            WebDriverWait(driver, 5).until(
                ec.element_to_be_clickable((By.XPATH, '//*[@title="Avanti"]'))
            ).click()
            logger.info("Pulsante 'Avanti' cliccato.")
        except Exception:
            logger.warning("Pulsante 'Avanti' non trovato. Procedo al passaggio successivo.")

        # Invio della query
        query = f"Estrai le seguenti informazioni, in formato json su una sola riga: {concat_fields(info_to_extract)} dal seguente testo non formattato: {input_text}"
        logger.info("Invio della query a Copilot.")
        try:
            question_box = WebDriverWait(driver, 10).until(
                ec.presence_of_element_located((By.ID, "userInput"))
            )
            question_box.send_keys(query, Keys.RETURN)
            logger.info("Query inviata con successo.")
        except Exception:
            logger.error("Casella di input per la query non trovata. Interazione fallita.")
            return None

        # Attesa della risposta
        logger.info("Attesa della risposta da Copilot.")
        time.sleep(10)
        response_element = WebDriverWait(driver, 10).until(
            ec.presence_of_element_located((By.XPATH, '//*[@class="text-sm font-ligatures-none"]'))
        )
        logger.info("Risposta ricevuta da Copilot.")

        response_text = response_element.text
        logger.info(response_text)

        return response_text

    except Exception as e:
        # In caso di errore durante l'interazione, viene loggato il messaggio di errore
        logger.error(f"Si è verificato un errore durante l'interazione con Copilot: {e}")
    finally:
        # Garantiamo che il driver venga sempre chiuso, anche se si è verificato un errore
        if driver:
            try:
                logger.info("Chiusura del browser.")
                driver.quit()
            except Exception as e:
                logger.error(f"Errore durante la chiusura del browser: {e}")


def select_columns_from_df(columns, df):
    """
    Seleziona specifiche colonne da un DataFrame.

    Parameters:
        columns (list): Lista di nomi di colonne da selezionare.
        df (pd.DataFrame): DataFrame di input.

    Returns:
        pd.DataFrame: DataFrame con solo le colonne selezionate.
    """
    if not isinstance(columns, list):
        raise ValueError("Il parametro 'columns' deve essere una lista di stringhe.")

    if not all(isinstance(col, str) for col in columns):
        raise ValueError("Tutti gli elementi della lista 'columns' devono essere stringhe.")

    missing_columns = [col for col in columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Le seguenti colonne non sono presenti nel DataFrame: {missing_columns}")

    return df[columns]


# Funzione per caricare i file già elaborati dal checkpoint
def load_processed_files(checkpoint_file):
    """
    Carica l'elenco dei file già elaborati dal file di checkpoint.

    Parameters:
        checkpoint_file (str): Percorso del file di checkpoint.

    Returns:
        set: Insieme dei nomi dei file già elaborati.
    """
    if os.path.exists(checkpoint_file):
        logger.info(f"Caricamento file elaborati dal checkpoint: {checkpoint_file}")
        with open(checkpoint_file, 'r') as f:
            processed_files = set(f.read().splitlines())
            logger.info(f"File elaborati caricati: {len(processed_files)} file trovati.")
            return processed_files
    else:
        logger.info(f"Nessun file di checkpoint trovato: {checkpoint_file}. Restituisco un set vuoto.")
        return set()


# Funzione per aggiornare il checkpoint con i nuovi file elaborati
def update_checkpoint(checkpoint_file, new_files):
    """
    Aggiorna il file di checkpoint aggiungendo i nuovi file elaborati.

    Parameters:
        checkpoint_file (str): Percorso del file di checkpoint.
        new_files: Insieme dei nuovi file elaborati da aggiungere.
    """
    logger.info(f"Aggiornamento del checkpoint con {len(new_files)} nuovi file.")
    with open(checkpoint_file, 'a') as f:
        for file in new_files:
            f.write(file + "\n")
    logger.info(f"Checkpoint aggiornato con successo: {checkpoint_file}")


# Cast delle colonne con gestione degli errori
def cast_columns_with_defaults(df, schema, default_values):
    # Fai una copia esplicita del DataFrame
    df = df.copy()

    for col, dtype in schema.items():
        if col in df.columns:
            try:
                # Prova a castare la colonna al tipo definito nello schema
                df.loc[:, col] = df[col].astype(dtype)
                logger.info(f"Colonna '{col}' castata con successo al tipo {dtype}.")
            except Exception as e:
                # In caso di errore, sostituisci con il valore predefinito
                logger.warning(f"Errore nel cast della colonna '{col}' al tipo {dtype}: {e}")
                default_value = default_values.get(dtype, None)
                df[col] = default_value
                logger.info(f"Colonna '{col}' riempita con valore predefinito: {default_value}.")
        else:
            # Se la colonna non esiste, aggiungila con il valore predefinito
            default_value = default_values.get(dtype, None)
            df[col] = default_value
            logger.info(f"Colonna '{col}' aggiunta al DataFrame con valore predefinito: {default_value}.")
    return df


def save_to_excel_in_append_mode(file_path, new_data):
    """
    Salva i dati nel file Excel in modalità append.

    Parametri:
    file_path (str): Percorso del file Excel.
    new_data (DataFrame): Il nuovo DataFrame da aggiungere.
    """
    if not os.path.exists(file_path):
        # Se il file non esiste, crea un nuovo file con i dati
        new_data.to_excel(file_path, index=False, engine='openpyxl')
        logger.info(f"File Excel creato: {file_path}")
    else:
        # Carica il file Excel esistente
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Ottieni l'indice della prima riga vuota
        start_row = sheet.max_row + 1

        # Scrivi i nuovi dati sotto quelli esistenti
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            new_data.to_excel(writer, index=False, header=False, startrow=start_row - 1)
        logger.info(f"Nuovi dati aggiunti al file Excel: {file_path}")


def upload_file_as_google_sheet(file_path, folder_id):
    """
    Carica un file su Google Drive come Google Fogli. Se esiste già, lo sovrascrive.

    Args:
        file_path (str): Percorso del file da caricare.
        folder_id (str): ID della cartella di Google Drive.
    """
    from os.path import splitext

    # Autenticazione
    credentials = service_account.Credentials.from_service_account_file(
        'resources/google/quiet-dimension-421414-f1378bc8723b.json',
        scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
    )
    drive_service = build('drive', 'v3', credentials=credentials)
    sheets_service = build('sheets', 'v4', credentials=credentials)

    # Ottieni il nome del file senza estensione
    file_name = file_path.split('/')[-1]
    base_name, _ = splitext(file_name)

    # Cerca se il file esiste già nella cartella (senza considerare l'estensione)
    query = f"'{folder_id}' in parents and name contains '{base_name}' and trashed=false"
    response = drive_service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = response.get('files', [])

    # Se il file esiste, ottieni il suo ID
    file_id = None
    if files:
        file_id = files[0]['id']
        logger.info(f"File esistente trovato con ID: {file_id}")

    # Leggi il file Excel con pandas
    df = pd.read_excel(file_path)

    # Converte le colonne di tipo datetime in stringhe
    for col in df.select_dtypes(include=['datetime64[ns]']).columns:
        df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')

    # Converte il DataFrame in una lista di liste (formato richiesto da Google Sheets)
    data = df.values.tolist()

    # Se il file esiste, sovrascrivi il contenuto
    if file_id:
        try:
            # Sovrascrivi tutto il contenuto del foglio
            update_request = sheets_service.spreadsheets().values().update(
                spreadsheetId=file_id,
                range="Sheet1!A1",  # Puoi specificare un altro intervallo se necessario
                body={"values": data},  # Aggiungi qui i dati che vuoi sovrascrivere
                valueInputOption="RAW"
            )
            update_request.execute()

            logger.info(f"Contenuto del file '{file_name}' sovrascritto.")
        except HttpError as err:
            logger.error(f"Errore nell'aggiornare il file: {err}")
    else:
        # Se il file non esiste, carica come nuovo
        file_metadata = {
            'name': file_name,
            'parents': [folder_id],
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        logger.info(f"File '{file_name}' caricato come Google Fogli con ID: {uploaded_file.get('id')}")


def create_database_and_table(host, port, user, password, db_name, new_db_name, table_name, table_schema):
    try:
        # Connessione al database principale
        conn = psycopg2.connect(
            host=host, port=port, user=user, password=password, dbname=db_name
        )
        conn.autocommit = True
        cursor = conn.cursor()

        # Creazione del database se non esiste
        cursor.execute(f"SELECT 1 FROM pg_database WHERE datname = '{new_db_name}'")
        if cursor.fetchone() is None:
            cursor.execute(f"CREATE DATABASE {new_db_name}")
            logger.info(f"Database '{new_db_name}' creato con successo.")
        else:
            logger.info(f"Database '{new_db_name}' già esistente.")

        cursor.close()
        conn.close()

        # Connessione al nuovo database per creare la tabella
        conn = psycopg2.connect(
            host=host, port=port, user=user, password=password, dbname=new_db_name
        )
        cursor = conn.cursor()

        # Creazione della tabella se non esiste
        cursor.execute(table_schema)
        conn.commit()
        logger.info(f"Tabella '{table_name}' pronta all'uso.")

        cursor.close()
        conn.close()

    except Exception as e:
        logger.info("Errore:", e)


def write_excel_to_table(host, port, user, password, db_name, table_name, file_path, insert_query):
    try:
        # Connessione al database
        conn = psycopg2.connect(
            host=host, port=port, user=user, password=password, dbname=db_name
        )
        cursor = conn.cursor()

        # Leggi l'Excel con Pandas
        df = pd.read_excel(file_path)

        # Overwrite dei dati nella tabella
        cursor.execute(f"TRUNCATE TABLE {table_name}")
        for _, row in df.iterrows():
            cursor.execute(insert_query, tuple(row))

        conn.commit()
        logger.info(f"Dati caricati con successo nella tabella '{table_name}'.")

        cursor.close()
        conn.close()

    except Exception as e:
        logger.error("Errore:", e)


def upload_to_airtable_from_excel(personal_token, base_id, table_name, excel_file_path):
    """
    Carica dati in una tabella di Airtable a partire da un file Excel.

    Args:
        personal_token (str): Il Personal Access Token di Airtable.
        base_id (str): L'ID della base Airtable.
        table_name (str): Il nome della tabella in cui caricare i dati.
        excel_file_path (str): Percorso del file Excel contenente i dati.

    Returns:
        dict: La risposta dell'API di Airtable o None in caso di errore.
    """
    # Validazione del percorso del file Excel
    if not excel_file_path.endswith(('.xls', '.xlsx')):
        logger.error("Errore: Il file specificato non è un file Excel valido.")
        return None

    # Leggi il file Excel e converti in una lista di dizionari
    try:
        df = pd.read_excel(excel_file_path)

        if df.empty:
            logger.error("Errore: Il file Excel è vuoto.")
            return None

        # Converti tutte le colonne datetime in stringhe ISO 8601
        for column in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[column]):
                df[column] = df[column].apply(lambda x: x.isoformat() if pd.notnull(x) else None)

        records = df.to_dict(orient="records")
    except Exception as e:
        logger.error(f"Errore nella lettura del file Excel: {e}")
        return None

    # Endpoint API di Airtable per la tabella specificata
    url = f"https://api.airtable.com/v0/{base_id}/{table_name}"

    # Intestazioni per l'autenticazione e il tipo di contenuto
    headers = {
        "Authorization": f"Bearer {personal_token}",
        "Content-Type": "application/json"
    }

    # Prepara i dati nel formato richiesto da Airtable
    payload = {
        "records": [{"fields": record} for record in records]
    }

    try:
        # Esegue la richiesta POST per caricare i dati
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()  # Solleva un'eccezione per eventuali errori HTTP

        # Restituisce la risposta in formato JSON
        logger.info("Dati caricati con successo su AirTable!")
        return response.json()
    except requests.exceptions.RequestException as req_err:
        logger.error(f"Errore nella richiesta: {req_err}")
    except Exception as e:
        logger.error(f"Errore imprevisto: {e}")

    return None
